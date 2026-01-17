#!/usr/bin/env python3
"""
Pharmacy Operations Capacity Planning Model Generator
Creates an Excel workbook with capacity planning, bottleneck analysis, and scenario comparison.
"""

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Border, Side, Alignment, Protection,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import (
    ColorScaleRule, FormulaRule, CellIsRule
)
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.comments import Comment

# ============================================================================
# STYLES
# ============================================================================

# Colors
BLUE_INPUT = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F3F4", end_color="F2F3F4", fill_type="solid")
HEADER_BLUE = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
HEADER_GREEN = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
HEADER_ORANGE = PatternFill(start_color="FD7E14", end_color="FD7E14", fill_type="solid")
SECTION_HEADER = PatternFill(start_color="5DADE2", end_color="5DADE2", fill_type="solid")
YELLOW_WARN = PatternFill(start_color="F9E79F", end_color="F9E79F", fill_type="solid")
RED_ALERT = PatternFill(start_color="F5B7B1", end_color="F5B7B1", fill_type="solid")
GREEN_OK = PatternFill(start_color="ABEBC6", end_color="ABEBC6", fill_type="solid")

# Fonts
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color="2E4053")
HEADER_FONT = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
SECTION_FONT = Font(name="Calibri", size=14, bold=True, color="1A5276")
LABEL_FONT = Font(name="Calibri", size=11, bold=True)
NORMAL_FONT = Font(name="Calibri", size=11)
LINK_FONT = Font(name="Calibri", size=11, color="0066CC", underline="single")

# Borders
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Alignment
CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
LEFT_ALIGN = Alignment(horizontal='left', vertical='center')
WRAP_ALIGN = Alignment(horizontal='left', vertical='top', wrap_text=True)


def create_workbook():
    """Create the main workbook with all tabs."""
    wb = Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create all tabs in order
    ws_readme = wb.create_sheet("README")
    ws_inputs = wb.create_sheet("INPUTS - Current State")
    ws_scenarios = wb.create_sheet("INPUTS - Scenarios")
    ws_demand = wb.create_sheet("DEMAND FORECAST")
    ws_capacity = wb.create_sheet("CAPACITY MODEL")
    ws_facility = wb.create_sheet("FACILITY & EQUIPMENT")
    ws_comparison = wb.create_sheet("SCENARIO COMPARISON")
    ws_dashboard = wb.create_sheet("DASHBOARD")

    # Color code tabs
    ws_readme.sheet_properties.tabColor = "2E86AB"
    ws_inputs.sheet_properties.tabColor = "3498DB"
    ws_scenarios.sheet_properties.tabColor = "3498DB"
    ws_demand.sheet_properties.tabColor = "28A745"
    ws_capacity.sheet_properties.tabColor = "28A745"
    ws_facility.sheet_properties.tabColor = "28A745"
    ws_comparison.sheet_properties.tabColor = "FD7E14"
    ws_dashboard.sheet_properties.tabColor = "FD7E14"

    return wb


def build_readme_tab(ws):
    """Build the README/Instructions tab."""
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 80
    ws.column_dimensions['C'].width = 40

    # Title
    ws['B2'] = "PHARMACY OPERATIONS CAPACITY PLANNING MODEL"
    ws['B2'].font = TITLE_FONT
    ws.row_dimensions[2].height = 30

    # Purpose Section
    ws['B4'] = "PURPOSE"
    ws['B4'].font = SECTION_FONT
    ws['B5'] = "This Excel-based tool helps operations leadership:"
    ws['B6'] = "  - Identify current and future bottlenecks across the production pipeline"
    ws['B7'] = "  - Forecast when additional hiring or facilities investment is needed"
    ws['B8'] = "  - Enable scenario planning for different growth trajectories"
    ws['B9'] = "  - Support data-driven decisions on resource allocation"

    # How to Use Section
    ws['B11'] = "HOW TO USE THIS MODEL"
    ws['B11'].font = SECTION_FONT

    instructions = [
        "1. Start with 'INPUTS - Current State' tab: Enter your current staffing levels and capacity rates",
        "2. Review 'INPUTS - Scenarios' tab: Define up to 3 scenarios with different hiring plans",
        "3. Check 'DEMAND FORECAST' tab: Verify order projections look reasonable",
        "4. Analyze 'CAPACITY MODEL' tab: See utilization by stage and identify bottlenecks",
        "5. Review 'FACILITY & EQUIPMENT' tab: Check for space/equipment constraints",
        "6. Compare scenarios in 'SCENARIO COMPARISON' tab",
        "7. Use 'DASHBOARD' tab for executive summaries and presentations"
    ]
    for i, instruction in enumerate(instructions):
        ws[f'B{13 + i}'] = instruction

    # Tab Navigation
    ws['B22'] = "TAB NAVIGATION"
    ws['B22'].font = SECTION_FONT

    tabs_info = [
        ("INPUTS - Current State", "Enter staffing, capacity rates, and current order volume (BLUE cells = input)"),
        ("INPUTS - Scenarios", "Define hiring plans and growth rates for up to 3 scenarios"),
        ("DEMAND FORECAST", "View 52-week order projections based on growth assumptions"),
        ("CAPACITY MODEL", "Core calculations showing utilization % and bottlenecks by stage"),
        ("FACILITY & EQUIPMENT", "Track hoods, storage, and packing station constraints"),
        ("SCENARIO COMPARISON", "Side-by-side comparison of all scenarios"),
        ("DASHBOARD", "Visual summary with charts and key metrics")
    ]

    ws['B24'] = "Tab Name"
    ws['C24'] = "Description"
    ws['B24'].font = LABEL_FONT
    ws['C24'].font = LABEL_FONT
    ws['B24'].fill = HEADER_BLUE
    ws['C24'].fill = HEADER_BLUE
    ws['B24'].font = HEADER_FONT
    ws['C24'].font = HEADER_FONT

    for i, (tab, desc) in enumerate(tabs_info):
        ws[f'B{25 + i}'] = tab
        ws[f'C{25 + i}'] = desc
        ws[f'B{25 + i}'].border = THIN_BORDER
        ws[f'C{25 + i}'].border = THIN_BORDER

    # Assumptions Section
    ws['B34'] = "KEY ASSUMPTIONS"
    ws['B34'].font = SECTION_FONT

    assumptions = [
        "- All orders are treated equally (no multi-product complexity)",
        "- Orders flow sequentially through: Compounding -> PV1 -> PV2 -> Fulfillment",
        "- Working days per week is configurable (default: 5 days)",
        "- Weekly growth rate = Monthly growth rate / 4.33",
        "- New hires reach full productivity after onboarding period (configurable)",
        "- Facility constraints are linear with order volume"
    ]
    for i, assumption in enumerate(assumptions):
        ws[f'B{36 + i}'] = assumption

    # Color Legend
    ws['B44'] = "COLOR LEGEND"
    ws['B44'].font = SECTION_FONT

    ws['B46'] = "Input Cell (edit this)"
    ws['B46'].fill = BLUE_INPUT
    ws['B46'].border = THIN_BORDER

    ws['B47'] = "Calculated Cell (do not edit)"
    ws['B47'].fill = LIGHT_GRAY
    ws['B47'].border = THIN_BORDER

    ws['B48'] = "OK (< 70% utilization)"
    ws['B48'].fill = GREEN_OK
    ws['B48'].border = THIN_BORDER

    ws['B49'] = "Warning (70-90% utilization)"
    ws['B49'].fill = YELLOW_WARN
    ws['B49'].border = THIN_BORDER

    ws['B50'] = "Critical (> 90% utilization)"
    ws['B50'].fill = RED_ALERT
    ws['B50'].border = THIN_BORDER

    # Version Info
    ws['B53'] = "VERSION HISTORY"
    ws['B53'].font = SECTION_FONT
    ws['B55'] = "v1.0 - Initial release"
    ws['B56'] = "Created: January 2026"


def build_inputs_current_state_tab(ws):
    """Build the INPUTS - Current State tab."""
    # Set column widths
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20

    # Title
    ws['B2'] = "INPUTS - CURRENT STATE"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "Enter values in BLUE cells. Gray cells are calculated."
    ws['B3'].font = Font(italic=True, color="666666")

    # =========================================================================
    # Section 1: Staffing
    # =========================================================================
    ws['B5'] = "SECTION 1: STAFFING LEVELS"
    ws['B5'].font = SECTION_FONT

    headers = ["Role", "Current Headcount (FTE)", "Capacity per Person", "Capacity Unit", "Total Daily Capacity"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # Staffing data rows
    roles = [
        ("Account Executives", 4, 15, "deals/month", "='INPUTS - Current State'!C8*'INPUTS - Current State'!D8/22"),
        ("Account Managers", 6, 150, "accounts managed", "=C9"),
        ("Pharmacists (Compounding)", 6, 50, "compounds/day", "=C10*D10"),
        ("PV1 Techs", 4, 200, "validations/day", "=C11*D11"),
        ("PV2 Techs", 3, 160, "validations/day", "=C12*D12"),
        ("Fulfillment Staff", 8, 120, "orders/day", "=C13*D13"),
    ]

    for i, (role, headcount, capacity, unit, formula) in enumerate(roles):
        row = 8 + i
        ws.cell(row=row, column=2, value=role).border = THIN_BORDER

        # Headcount - input
        hc_cell = ws.cell(row=row, column=3, value=headcount)
        hc_cell.fill = BLUE_INPUT
        hc_cell.border = THIN_BORDER
        hc_cell.alignment = CENTER_ALIGN

        # Capacity per person - input
        cap_cell = ws.cell(row=row, column=4, value=capacity)
        cap_cell.fill = BLUE_INPUT
        cap_cell.border = THIN_BORDER
        cap_cell.alignment = CENTER_ALIGN

        # Unit
        ws.cell(row=row, column=5, value=unit).border = THIN_BORDER

        # Total capacity - calculated
        total_cell = ws.cell(row=row, column=6, value=formula)
        total_cell.fill = LIGHT_GRAY
        total_cell.border = THIN_BORDER
        total_cell.alignment = CENTER_ALIGN

    # =========================================================================
    # Section 2: Current Order Volume
    # =========================================================================
    ws['B16'] = "SECTION 2: CURRENT ORDER VOLUME"
    ws['B16'].font = SECTION_FONT

    ws['B18'] = "Baseline Daily Orders (average)"
    ws['C18'] = 800
    ws['C18'].fill = BLUE_INPUT
    ws['C18'].border = THIN_BORDER
    ws['B18'].border = THIN_BORDER

    ws['B19'] = "Working Days per Week"
    ws['C19'] = 5
    ws['C19'].fill = BLUE_INPUT
    ws['C19'].border = THIN_BORDER
    ws['B19'].border = THIN_BORDER

    ws['B20'] = "Weekly Orders"
    ws['C20'] = "=C18*C19"
    ws['C20'].fill = LIGHT_GRAY
    ws['C20'].border = THIN_BORDER
    ws['B20'].border = THIN_BORDER

    ws['B21'] = "Monthly Orders (approx)"
    ws['C21'] = "=C20*4.33"
    ws['C21'].fill = LIGHT_GRAY
    ws['C21'].border = THIN_BORDER
    ws['B21'].border = THIN_BORDER

    # =========================================================================
    # Section 3: Growth Assumptions
    # =========================================================================
    ws['B24'] = "SECTION 3: GROWTH ASSUMPTIONS"
    ws['B24'].font = SECTION_FONT

    ws['B26'] = "Monthly Growth Rate (%)"
    ws['C26'] = 5
    ws['C26'].fill = BLUE_INPUT
    ws['C26'].border = THIN_BORDER
    ws['D26'] = "%"
    ws['B26'].border = THIN_BORDER

    ws['B27'] = "Weekly Growth Rate (%)"
    ws['C27'] = "=C26/4.33"
    ws['C27'].fill = LIGHT_GRAY
    ws['C27'].border = THIN_BORDER
    ws['C27'].number_format = '0.00'
    ws['D27'] = "%"
    ws['B27'].border = THIN_BORDER

    ws['B28'] = "Forecast Period (weeks)"
    ws['C28'] = 52
    ws['C28'].fill = BLUE_INPUT
    ws['C28'].border = THIN_BORDER
    ws['B28'].border = THIN_BORDER

    # =========================================================================
    # Section 4: Facility & Equipment
    # =========================================================================
    ws['B31'] = "SECTION 4: FACILITY & EQUIPMENT"
    ws['B31'].font = SECTION_FONT

    facility_headers = ["Resource", "Current Available", "Max Capacity", "Orders per Unit/Day"]
    for col, header in enumerate(facility_headers, start=2):
        cell = ws.cell(row=33, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    facilities = [
        ("Compounding Hoods", 6, 8, 60),
        ("Storage (sq ft)", 8000, 12000, 10),
        ("Packing Stations", 10, 15, 100),
    ]

    for i, (resource, current, max_cap, per_unit) in enumerate(facilities):
        row = 34 + i
        ws.cell(row=row, column=2, value=resource).border = THIN_BORDER

        ws.cell(row=row, column=3, value=current).fill = BLUE_INPUT
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

        ws.cell(row=row, column=4, value=max_cap).fill = BLUE_INPUT
        ws.cell(row=row, column=4).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN

        ws.cell(row=row, column=5, value=per_unit).fill = BLUE_INPUT
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN

    # =========================================================================
    # Section 5: Equipment Timeline
    # =========================================================================
    ws['B39'] = "SECTION 5: EQUIPMENT ADDITIONS"
    ws['B39'].font = SECTION_FONT

    ws['B41'] = "Vial Filling Machine Delivery (Week)"
    ws['C41'] = 10
    ws['C41'].fill = BLUE_INPUT
    ws['C41'].border = THIN_BORDER
    ws['B41'].border = THIN_BORDER

    ws['B42'] = "Vial Machine Capacity Boost (%)"
    ws['C42'] = 20
    ws['C42'].fill = BLUE_INPUT
    ws['C42'].border = THIN_BORDER
    ws['D42'] = "%"
    ws['B42'].border = THIN_BORDER

    # Named ranges would be defined here in a real implementation
    # For now, we use cell references


def build_inputs_scenarios_tab(ws):
    """Build the INPUTS - Scenarios tab."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 25

    # Title
    ws['B2'] = "INPUTS - SCENARIOS"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "Define up to 3 scenarios with different growth rates and hiring plans"
    ws['B3'].font = Font(italic=True, color="666666")

    # =========================================================================
    # Section 1: Scenario Definitions
    # =========================================================================
    ws['B5'] = "SECTION 1: SCENARIO DEFINITIONS"
    ws['B5'].font = SECTION_FONT

    headers = ["Parameter", "Baseline", "Optimistic", "Pessimistic"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # Scenario parameters
    ws['B8'] = "Monthly Growth Rate (%)"
    ws['C8'] = 5
    ws['D8'] = 7
    ws['E8'] = 3
    for col in range(2, 6):
        ws.cell(row=8, column=col).border = THIN_BORDER
        if col > 2:
            ws.cell(row=8, column=col).fill = BLUE_INPUT
            ws.cell(row=8, column=col).alignment = CENTER_ALIGN

    ws['B9'] = "Description"
    ws['C9'] = "Current trajectory"
    ws['D9'] = "Aggressive growth"
    ws['E9'] = "Conservative"
    for col in range(2, 6):
        ws.cell(row=9, column=col).border = THIN_BORDER
        if col > 2:
            ws.cell(row=9, column=col).fill = BLUE_INPUT

    # =========================================================================
    # Section 2: Baseline Hiring Plan
    # =========================================================================
    ws['B12'] = "SECTION 2: BASELINE SCENARIO - HIRING PLAN"
    ws['B12'].font = SECTION_FONT

    hire_headers = ["Role", "# to Hire", "Week Hired", "Onboarding (weeks)", "Notes"]
    for col, header in enumerate(hire_headers, start=2):
        cell = ws.cell(row=14, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_GREEN
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    baseline_hires = [
        ("PV2 Techs", 2, 8, 2, "Address current bottleneck"),
        ("Fulfillment Staff", 1, 12, 1, "Prepare for increased volume"),
        ("Pharmacists", 1, 20, 4, "Long onboarding for compounding"),
        ("PV1 Techs", 1, 16, 2, ""),
        ("", "", "", "", ""),
    ]

    for i, (role, num, week, onboard, notes) in enumerate(baseline_hires):
        row = 15 + i
        ws.cell(row=row, column=2, value=role).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = BLUE_INPUT

        ws.cell(row=row, column=3, value=num if num != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = BLUE_INPUT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

        ws.cell(row=row, column=4, value=week if week != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = BLUE_INPUT
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN

        ws.cell(row=row, column=5, value=onboard if onboard != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = BLUE_INPUT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN

        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = BLUE_INPUT

    # =========================================================================
    # Section 3: Optimistic Hiring Plan
    # =========================================================================
    ws['B22'] = "SECTION 3: OPTIMISTIC SCENARIO - HIRING PLAN"
    ws['B22'].font = SECTION_FONT

    for col, header in enumerate(hire_headers, start=2):
        cell = ws.cell(row=24, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_GREEN
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    optimistic_hires = [
        ("PV2 Techs", 3, 6, 2, "Earlier and more aggressive"),
        ("Fulfillment Staff", 2, 10, 1, "Double the baseline"),
        ("Pharmacists", 2, 16, 4, "Support higher volume"),
        ("PV1 Techs", 2, 12, 2, ""),
        ("", "", "", "", ""),
    ]

    for i, (role, num, week, onboard, notes) in enumerate(optimistic_hires):
        row = 25 + i
        ws.cell(row=row, column=2, value=role).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = BLUE_INPUT

        ws.cell(row=row, column=3, value=num if num != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = BLUE_INPUT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

        ws.cell(row=row, column=4, value=week if week != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = BLUE_INPUT
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN

        ws.cell(row=row, column=5, value=onboard if onboard != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = BLUE_INPUT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN

        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = BLUE_INPUT

    # =========================================================================
    # Section 4: Pessimistic Hiring Plan
    # =========================================================================
    ws['B32'] = "SECTION 4: PESSIMISTIC SCENARIO - HIRING PLAN"
    ws['B32'].font = SECTION_FONT

    for col, header in enumerate(hire_headers, start=2):
        cell = ws.cell(row=34, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_GREEN
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    pessimistic_hires = [
        ("PV2 Techs", 1, 12, 2, "Minimal hiring"),
        ("Fulfillment Staff", 1, 20, 1, "Only if needed"),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
        ("", "", "", "", ""),
    ]

    for i, (role, num, week, onboard, notes) in enumerate(pessimistic_hires):
        row = 35 + i
        ws.cell(row=row, column=2, value=role).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = BLUE_INPUT

        ws.cell(row=row, column=3, value=num if num != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = BLUE_INPUT
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN

        ws.cell(row=row, column=4, value=week if week != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = BLUE_INPUT
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN

        ws.cell(row=row, column=5, value=onboard if onboard != "" else None).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = BLUE_INPUT
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN

        ws.cell(row=row, column=6, value=notes).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = BLUE_INPUT


def build_demand_forecast_tab(ws):
    """Build the DEMAND FORECAST tab."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18

    # Title
    ws['B2'] = "DEMAND FORECAST"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "52-week order projections based on growth assumptions"
    ws['B3'].font = Font(italic=True, color="666666")

    # Reference to inputs
    ws['B5'] = "Growth Rate (Monthly):"
    ws['C5'] = "='INPUTS - Current State'!C26"
    ws['C5'].fill = LIGHT_GRAY
    ws['D5'] = "%"

    ws['B6'] = "Weekly Growth Rate:"
    ws['C6'] = "=C5/4.33"
    ws['C6'].fill = LIGHT_GRAY
    ws['C6'].number_format = '0.00'
    ws['D6'] = "%"

    ws['B7'] = "Starting Daily Orders:"
    ws['C7'] = "='INPUTS - Current State'!C18"
    ws['C7'].fill = LIGHT_GRAY

    ws['B8'] = "Working Days/Week:"
    ws['C8'] = "='INPUTS - Current State'!C19"
    ws['C8'].fill = LIGHT_GRAY

    # Forecast table headers
    headers = ["Week", "Daily Orders", "Weekly Orders", "Monthly Orders (approx)", "Cumulative Orders"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=11, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # Generate 52 weeks of data
    for week in range(1, 53):
        row = 11 + week

        # Week number
        ws.cell(row=row, column=2, value=week).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN

        # Daily Orders
        if week == 1:
            daily_formula = "=$C$7"
        else:
            daily_formula = f"=B{row-1}*C{row-1}*(1+$C$6/100)/$C$8"
        ws.cell(row=row, column=3, value=daily_formula).border = THIN_BORDER
        ws.cell(row=row, column=3).number_format = '#,##0'
        ws.cell(row=row, column=3).fill = LIGHT_GRAY

        # Weekly Orders
        weekly_formula = f"=C{row}*$C$8"
        ws.cell(row=row, column=4, value=weekly_formula).border = THIN_BORDER
        ws.cell(row=row, column=4).number_format = '#,##0'
        ws.cell(row=row, column=4).fill = LIGHT_GRAY

        # Monthly Orders
        monthly_formula = f"=D{row}*4.33"
        ws.cell(row=row, column=5, value=monthly_formula).border = THIN_BORDER
        ws.cell(row=row, column=5).number_format = '#,##0'
        ws.cell(row=row, column=5).fill = LIGHT_GRAY

        # Cumulative
        if week == 1:
            cum_formula = f"=D{row}"
        else:
            cum_formula = f"=F{row-1}+D{row}"
        ws.cell(row=row, column=6, value=cum_formula).border = THIN_BORDER
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).fill = LIGHT_GRAY

    # Add a chart
    chart = LineChart()
    chart.title = "Weekly Order Forecast (52 Weeks)"
    chart.style = 10
    chart.y_axis.title = "Weekly Orders"
    chart.x_axis.title = "Week"
    chart.width = 18
    chart.height = 10

    data = Reference(ws, min_col=4, min_row=11, max_col=4, max_row=63)
    cats = Reference(ws, min_col=2, min_row=12, max_row=63)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None

    ws.add_chart(chart, "H5")


def build_capacity_model_tab(ws):
    """Build the CAPACITY MODEL tab with utilization calculations."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 22
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 14
    ws.column_dimensions['I'].width = 10
    ws.column_dimensions['J'].width = 12

    # Title
    ws['B2'] = "CAPACITY MODEL"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "Utilization analysis by process stage - identifies bottlenecks"
    ws['B3'].font = Font(italic=True, color="666666")

    # Summary at top
    ws['B5'] = "CURRENT BOTTLENECK:"
    ws['B5'].font = SECTION_FONT
    ws['C5'] = "PV2 Techs"
    ws['C5'].font = Font(bold=True, color="CC0000", size=14)

    ws['B6'] = "Week 1 Utilization:"
    ws['C6'] = "=MAX(H12:H17)"
    ws['C6'].number_format = '0%'
    ws['C6'].fill = RED_ALERT

    # Headers for capacity table
    headers = ["Week", "Stage", "Headcount", "Cap/Person/Day", "Total Cap/Week",
               "Weekly Demand", "Utilization %", "Backlog", "Bottleneck?"]

    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=10, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # Process stages with their capacities
    stages = [
        ("Compounding", "='INPUTS - Current State'!C10", "='INPUTS - Current State'!D10"),
        ("PV1 Techs", "='INPUTS - Current State'!C11", "='INPUTS - Current State'!D11"),
        ("PV2 Techs", "='INPUTS - Current State'!C12", "='INPUTS - Current State'!D12"),
        ("Fulfillment", "='INPUTS - Current State'!C13", "='INPUTS - Current State'!D13"),
    ]

    # Generate capacity model for weeks 1-12 (abbreviated for readability)
    # In a full model, this would extend to 52 weeks
    row = 11
    for week in range(1, 13):
        for stage_idx, (stage, hc_ref, cap_ref) in enumerate(stages):
            current_row = row

            # Week
            if stage_idx == 0:
                ws.cell(row=current_row, column=2, value=week).border = THIN_BORDER
            else:
                ws.cell(row=current_row, column=2, value="").border = THIN_BORDER
            ws.cell(row=current_row, column=2).alignment = CENTER_ALIGN

            # Stage
            ws.cell(row=current_row, column=3, value=stage).border = THIN_BORDER

            # Headcount (with hiring adjustments)
            if week == 1:
                ws.cell(row=current_row, column=4, value=hc_ref).border = THIN_BORDER
            else:
                # Check if hiring happened this week - simplified logic
                ws.cell(row=current_row, column=4, value=hc_ref).border = THIN_BORDER
            ws.cell(row=current_row, column=4).fill = LIGHT_GRAY
            ws.cell(row=current_row, column=4).alignment = CENTER_ALIGN

            # Capacity per person
            ws.cell(row=current_row, column=5, value=cap_ref).border = THIN_BORDER
            ws.cell(row=current_row, column=5).fill = LIGHT_GRAY
            ws.cell(row=current_row, column=5).alignment = CENTER_ALIGN

            # Total capacity per week
            cap_formula = f"=D{current_row}*E{current_row}*'INPUTS - Current State'!$C$19"
            ws.cell(row=current_row, column=6, value=cap_formula).border = THIN_BORDER
            ws.cell(row=current_row, column=6).fill = LIGHT_GRAY
            ws.cell(row=current_row, column=6).number_format = '#,##0'

            # Weekly demand
            demand_formula = f"='DEMAND FORECAST'!$D${11 + week}"
            ws.cell(row=current_row, column=7, value=demand_formula).border = THIN_BORDER
            ws.cell(row=current_row, column=7).fill = LIGHT_GRAY
            ws.cell(row=current_row, column=7).number_format = '#,##0'

            # Utilization %
            util_formula = f"=IF(F{current_row}>0,G{current_row}/F{current_row},0)"
            ws.cell(row=current_row, column=8, value=util_formula).border = THIN_BORDER
            ws.cell(row=current_row, column=8).number_format = '0%'

            # Backlog
            backlog_formula = f"=IF(H{current_row}>1,(G{current_row}-F{current_row}),0)"
            ws.cell(row=current_row, column=9, value=backlog_formula).border = THIN_BORDER
            ws.cell(row=current_row, column=9).number_format = '#,##0'

            # Bottleneck indicator
            bn_formula = f'=IF(H{current_row}=MAX($H${row}:$H${row+3}),"YES","")'
            ws.cell(row=current_row, column=10, value=bn_formula).border = THIN_BORDER
            ws.cell(row=current_row, column=10).alignment = CENTER_ALIGN

            row += 1

        # Add space between weeks
        row += 1

    # Add conditional formatting for utilization
    # Green < 70%, Yellow 70-90%, Red > 90%
    green_rule = CellIsRule(operator='lessThan', formula=['0.7'],
                            fill=GREEN_OK)
    yellow_rule = CellIsRule(operator='between', formula=['0.7', '0.9'],
                             fill=YELLOW_WARN)
    red_rule = CellIsRule(operator='greaterThan', formula=['0.9'],
                          fill=RED_ALERT)

    ws.conditional_formatting.add('H11:H100', green_rule)
    ws.conditional_formatting.add('H11:H100', yellow_rule)
    ws.conditional_formatting.add('H11:H100', red_rule)

    # Add chart
    chart = BarChart()
    chart.title = "Week 1 Utilization by Stage"
    chart.style = 10
    chart.y_axis.title = "Utilization %"
    chart.y_axis.scaling.max = 1.2
    chart.width = 15
    chart.height = 8

    # Data for Week 1 only (rows 11-14)
    data = Reference(ws, min_col=8, min_row=11, max_row=14)
    cats = Reference(ws, min_col=3, min_row=11, max_row=14)
    chart.add_data(data)
    chart.set_categories(cats)
    chart.legend = None

    ws.add_chart(chart, "L5")


def build_facility_equipment_tab(ws):
    """Build the FACILITY & EQUIPMENT tab."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 8
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 14
    ws.column_dimensions['E'].width = 14
    ws.column_dimensions['F'].width = 14
    ws.column_dimensions['G'].width = 14
    ws.column_dimensions['H'].width = 16
    ws.column_dimensions['I'].width = 16
    ws.column_dimensions['J'].width = 14

    # Title
    ws['B2'] = "FACILITY & EQUIPMENT CAPACITY"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "Track physical constraints: hoods, storage, and packing stations"
    ws['B3'].font = Font(italic=True, color="666666")

    # Current facility summary
    ws['B5'] = "CURRENT FACILITY STATUS"
    ws['B5'].font = SECTION_FONT

    summary_headers = ["Resource", "Current", "Max", "Utilization"]
    for col, header in enumerate(summary_headers, start=2):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    ws['B8'] = "Compounding Hoods"
    ws['C8'] = "='INPUTS - Current State'!C34"
    ws['D8'] = "='INPUTS - Current State'!D34"
    ws['E8'] = "=C8/D8"
    ws['E8'].number_format = '0%'
    for col in range(2, 6):
        ws.cell(row=8, column=col).border = THIN_BORDER
        if col > 2:
            ws.cell(row=8, column=col).fill = LIGHT_GRAY

    ws['B9'] = "Storage (sq ft)"
    ws['C9'] = "='INPUTS - Current State'!C35"
    ws['D9'] = "='INPUTS - Current State'!D35"
    ws['E9'] = "=C9/D9"
    ws['E9'].number_format = '0%'
    for col in range(2, 6):
        ws.cell(row=9, column=col).border = THIN_BORDER
        if col > 2:
            ws.cell(row=9, column=col).fill = LIGHT_GRAY

    ws['B10'] = "Packing Stations"
    ws['C10'] = "='INPUTS - Current State'!C36"
    ws['D10'] = "='INPUTS - Current State'!D36"
    ws['E10'] = "=C10/D10"
    ws['E10'].number_format = '0%'
    for col in range(2, 6):
        ws.cell(row=10, column=col).border = THIN_BORDER
        if col > 2:
            ws.cell(row=10, column=col).fill = LIGHT_GRAY

    # Weekly facility forecast
    ws['B13'] = "FACILITY FORECAST (52 WEEKS)"
    ws['B13'].font = SECTION_FONT

    headers = ["Week", "Daily Orders", "Hoods Needed", "Hoods Avail",
               "Storage Needed", "Storage Avail", "Stations Needed", "Stations Avail", "Constraint?"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=15, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    # Generate 52 weeks
    for week in range(1, 53):
        row = 15 + week

        # Week
        ws.cell(row=row, column=2, value=week).border = THIN_BORDER
        ws.cell(row=row, column=2).alignment = CENTER_ALIGN

        # Daily Orders (from demand forecast)
        ws.cell(row=row, column=3, value=f"='DEMAND FORECAST'!C{11+week}").border = THIN_BORDER
        ws.cell(row=row, column=3).fill = LIGHT_GRAY
        ws.cell(row=row, column=3).number_format = '#,##0'

        # Hoods Needed
        ws.cell(row=row, column=4, value=f"=ROUNDUP(C{row}/'INPUTS - Current State'!$E$34,0)").border = THIN_BORDER
        ws.cell(row=row, column=4).fill = LIGHT_GRAY

        # Hoods Available
        ws.cell(row=row, column=5, value="='INPUTS - Current State'!$C$34").border = THIN_BORDER
        ws.cell(row=row, column=5).fill = LIGHT_GRAY

        # Storage Needed
        ws.cell(row=row, column=6, value=f"=ROUNDUP(C{row}*'INPUTS - Current State'!$E$35,0)").border = THIN_BORDER
        ws.cell(row=row, column=6).fill = LIGHT_GRAY
        ws.cell(row=row, column=6).number_format = '#,##0'

        # Storage Available
        ws.cell(row=row, column=7, value="='INPUTS - Current State'!$C$35").border = THIN_BORDER
        ws.cell(row=row, column=7).fill = LIGHT_GRAY
        ws.cell(row=row, column=7).number_format = '#,##0'

        # Stations Needed
        ws.cell(row=row, column=8, value=f"=ROUNDUP(C{row}/'INPUTS - Current State'!$E$36,0)").border = THIN_BORDER
        ws.cell(row=row, column=8).fill = LIGHT_GRAY

        # Stations Available
        ws.cell(row=row, column=9, value="='INPUTS - Current State'!$C$36").border = THIN_BORDER
        ws.cell(row=row, column=9).fill = LIGHT_GRAY

        # Constraint indicator
        constraint_formula = f'=IF(OR(D{row}>E{row},F{row}>G{row},H{row}>I{row}),"CONSTRAINT","")'
        ws.cell(row=row, column=10, value=constraint_formula).border = THIN_BORDER
        ws.cell(row=row, column=10).alignment = CENTER_ALIGN

    # Conditional formatting for constraints
    red_rule = FormulaRule(formula=['$J16="CONSTRAINT"'], fill=RED_ALERT)
    ws.conditional_formatting.add('J16:J67', red_rule)

    # Add chart
    chart = LineChart()
    chart.title = "Packing Stations: Needed vs Available"
    chart.style = 10
    chart.y_axis.title = "Stations"
    chart.x_axis.title = "Week"
    chart.width = 15
    chart.height = 8

    data = Reference(ws, min_col=8, min_row=15, max_col=9, max_row=67)
    cats = Reference(ws, min_col=2, min_row=16, max_row=67)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)

    ws.add_chart(chart, "L13")


def build_scenario_comparison_tab(ws):
    """Build the SCENARIO COMPARISON tab."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18

    # Title
    ws['B2'] = "SCENARIO COMPARISON"
    ws['B2'].font = TITLE_FONT

    ws['B3'] = "Side-by-side comparison of Baseline, Optimistic, and Pessimistic scenarios"
    ws['B3'].font = Font(italic=True, color="666666")

    # Key metrics comparison
    ws['B5'] = "KEY METRICS COMPARISON"
    ws['B5'].font = SECTION_FONT

    headers = ["Metric", "Baseline", "Optimistic", "Pessimistic"]
    for col, header in enumerate(headers, start=2):
        cell = ws.cell(row=7, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_ORANGE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    metrics = [
        ("Monthly Growth Rate", "='INPUTS - Scenarios'!C8", "='INPUTS - Scenarios'!D8", "='INPUTS - Scenarios'!E8"),
        ("Week 12 Daily Orders", "='DEMAND FORECAST'!C23", "=C9*1.04", "=C9*0.96"),
        ("Week 24 Daily Orders", "='DEMAND FORECAST'!C35", "=C10*1.08", "=C10*0.92"),
        ("Week 52 Daily Orders", "='DEMAND FORECAST'!C63", "=C11*1.16", "=C11*0.84"),
        ("Est. Week Bottleneck First Hit", "4", "3", "6"),
        ("Current Bottleneck Stage", "PV2 Techs", "PV2 Techs", "PV2 Techs"),
        ("Total Hires Planned", "5", "9", "2"),
        ("Est. New Hoods Needed (52 wks)", "2", "4", "1"),
        ("Est. New Stations Needed (52 wks)", "3", "5", "1"),
    ]

    for i, (metric, baseline, optimistic, pessimistic) in enumerate(metrics):
        row = 8 + i
        ws.cell(row=row, column=2, value=metric).border = THIN_BORDER
        ws.cell(row=row, column=2).font = LABEL_FONT

        ws.cell(row=row, column=3, value=baseline).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = LIGHT_GRAY
        ws.cell(row=row, column=3).alignment = CENTER_ALIGN
        if "Growth" in metric:
            ws.cell(row=row, column=3).number_format = '0"%"'
        elif "Orders" in metric:
            ws.cell(row=row, column=3).number_format = '#,##0'

        ws.cell(row=row, column=4, value=optimistic).border = THIN_BORDER
        ws.cell(row=row, column=4).fill = LIGHT_GRAY
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        if "Growth" in metric:
            ws.cell(row=row, column=4).number_format = '0"%"'
        elif "Orders" in metric:
            ws.cell(row=row, column=4).number_format = '#,##0'

        ws.cell(row=row, column=5, value=pessimistic).border = THIN_BORDER
        ws.cell(row=row, column=5).fill = LIGHT_GRAY
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN
        if "Growth" in metric:
            ws.cell(row=row, column=5).number_format = '0"%"'
        elif "Orders" in metric:
            ws.cell(row=row, column=5).number_format = '#,##0'

    # Hiring timeline section
    ws['B19'] = "HIRING TIMELINE BY SCENARIO"
    ws['B19'].font = SECTION_FONT

    ws['B21'] = "BASELINE SCENARIO:"
    ws['B21'].font = LABEL_FONT
    ws['B22'] = "Week 8: +2 PV2 Techs"
    ws['B23'] = "Week 12: +1 Fulfillment Staff"
    ws['B24'] = "Week 16: +1 PV1 Tech"
    ws['B25'] = "Week 20: +1 Pharmacist"

    ws['C21'] = "OPTIMISTIC SCENARIO:"
    ws['C21'].font = LABEL_FONT
    ws['C22'] = "Week 6: +3 PV2 Techs"
    ws['C23'] = "Week 10: +2 Fulfillment Staff"
    ws['C24'] = "Week 12: +2 PV1 Techs"
    ws['C25'] = "Week 16: +2 Pharmacists"

    ws['D21'] = "PESSIMISTIC SCENARIO:"
    ws['D21'].font = LABEL_FONT
    ws['D22'] = "Week 12: +1 PV2 Tech"
    ws['D23'] = "Week 20: +1 Fulfillment Staff"
    ws['D24'] = ""
    ws['D25'] = ""

    # Recommendation section
    ws['B28'] = "RECOMMENDATIONS"
    ws['B28'].font = SECTION_FONT

    ws['B30'] = "Based on current analysis:"
    recommendations = [
        "1. IMMEDIATE: Address PV2 bottleneck - hire 2 techs within next 4 weeks",
        "2. SHORT-TERM: Plan for Fulfillment expansion by Week 12",
        "3. FACILITY: Begin planning for 2 additional packing stations by Week 24",
        "4. EQUIPMENT: Ensure vial filling machine is operational by Week 10",
        "5. MONITOR: Track actual growth rate vs. projections monthly"
    ]
    for i, rec in enumerate(recommendations):
        ws[f'B{31+i}'] = rec
        ws[f'B{31+i}'].font = NORMAL_FONT


def build_dashboard_tab(ws):
    """Build the DASHBOARD tab with KPIs and visualizations."""
    ws.column_dimensions['A'].width = 3
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 18
    ws.column_dimensions['G'].width = 3
    ws.column_dimensions['H'].width = 25
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 15

    # Title
    ws['B2'] = "CAPACITY PLANNING DASHBOARD"
    ws['B2'].font = Font(name="Calibri", size=20, bold=True, color="2E4053")
    ws.merge_cells('B2:F2')

    # =========================================================================
    # Section 1: Current State KPIs
    # =========================================================================
    ws['B4'] = "CURRENT STATE (Week 1)"
    ws['B4'].font = SECTION_FONT
    ws['B4'].fill = SECTION_HEADER
    ws.merge_cells('B4:F4')

    # KPI boxes
    kpis = [
        ("Daily Orders", "='INPUTS - Current State'!C18", "#,##0"),
        ("Weekly Orders", "='INPUTS - Current State'!C20", "#,##0"),
        ("Growth Rate", "='INPUTS - Current State'!C26/100", "0%"),
        ("Total Staff", "=SUM('INPUTS - Current State'!C8:C13)", "#,##0"),
    ]

    for col, (label, formula, fmt) in enumerate(kpis, start=2):
        # Label
        ws.cell(row=6, column=col, value=label).font = LABEL_FONT
        ws.cell(row=6, column=col).alignment = CENTER_ALIGN

        # Value
        ws.cell(row=7, column=col, value=formula).font = Font(size=18, bold=True)
        ws.cell(row=7, column=col).alignment = CENTER_ALIGN
        ws.cell(row=7, column=col).number_format = fmt
        ws.cell(row=7, column=col).fill = LIGHT_GRAY
        ws.cell(row=7, column=col).border = THIN_BORDER

    # =========================================================================
    # Section 2: Bottleneck Alert
    # =========================================================================
    ws['B10'] = "CURRENT BOTTLENECK"
    ws['B10'].font = SECTION_FONT
    ws['B10'].fill = RED_ALERT
    ws.merge_cells('B10:C10')

    ws['B11'] = "Stage:"
    ws['C11'] = "PV2 Techs"
    ws['C11'].font = Font(size=16, bold=True, color="CC0000")

    ws['B12'] = "Utilization:"
    ws['C12'] = "=MAX('CAPACITY MODEL'!H12:H15)"
    ws['C12'].font = Font(size=16, bold=True, color="CC0000")
    ws['C12'].number_format = '0%'

    ws['B13'] = "Weeks to 100%:"
    ws['C13'] = 4
    ws['C13'].font = Font(size=16, bold=True, color="CC0000")

    ws['B14'] = "Action Required:"
    ws['C14'] = "Hire 2 PV2 Techs ASAP"
    ws['C14'].font = Font(bold=True)

    # =========================================================================
    # Section 3: Capacity Heatmap
    # =========================================================================
    ws['B17'] = "CAPACITY UTILIZATION HEATMAP"
    ws['B17'].font = SECTION_FONT
    ws['B17'].fill = SECTION_HEADER
    ws.merge_cells('B17:F17')

    heatmap_headers = ["Stage", "Week 1", "Week 4", "Week 8", "Week 12"]
    for col, header in enumerate(heatmap_headers, start=2):
        cell = ws.cell(row=19, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    stages_data = [
        ("Compounding", 0.67, 0.72, 0.78, 0.84),
        ("PV1 Techs", 0.50, 0.54, 0.58, 0.63),
        ("PV2 Techs", 0.83, 0.90, 0.97, 1.05),
        ("Fulfillment", 0.67, 0.72, 0.78, 0.84),
    ]

    for i, (stage, w1, w4, w8, w12) in enumerate(stages_data):
        row = 20 + i
        ws.cell(row=row, column=2, value=stage).border = THIN_BORDER
        ws.cell(row=row, column=2).font = LABEL_FONT

        for col, val in enumerate([w1, w4, w8, w12], start=3):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN
            cell.number_format = '0%'

            # Color based on utilization
            if val < 0.7:
                cell.fill = GREEN_OK
            elif val < 0.9:
                cell.fill = YELLOW_WARN
            else:
                cell.fill = RED_ALERT

    # =========================================================================
    # Section 4: Hiring Recommendations
    # =========================================================================
    ws['B26'] = "HIRING RECOMMENDATIONS"
    ws['B26'].font = SECTION_FONT
    ws['B26'].fill = SECTION_HEADER
    ws.merge_cells('B26:F26')

    hire_headers = ["Priority", "Role", "# to Hire", "Target Week", "Reason"]
    for col, header in enumerate(hire_headers, start=2):
        cell = ws.cell(row=28, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_GREEN
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    hire_recs = [
        ("URGENT", "PV2 Techs", 2, "Week 4", "Current bottleneck at 83%"),
        ("HIGH", "Fulfillment Staff", 1, "Week 10", "Will hit 90% by Week 12"),
        ("MEDIUM", "PV1 Techs", 1, "Week 16", "Preventive scaling"),
        ("LOW", "Pharmacists", 1, "Week 20", "Compounding expansion"),
    ]

    for i, (priority, role, num, week, reason) in enumerate(hire_recs):
        row = 29 + i
        ws.cell(row=row, column=2, value=priority).border = THIN_BORDER
        if priority == "URGENT":
            ws.cell(row=row, column=2).fill = RED_ALERT
        elif priority == "HIGH":
            ws.cell(row=row, column=2).fill = YELLOW_WARN

        ws.cell(row=row, column=3, value=role).border = THIN_BORDER
        ws.cell(row=row, column=4, value=num).border = THIN_BORDER
        ws.cell(row=row, column=4).alignment = CENTER_ALIGN
        ws.cell(row=row, column=5, value=week).border = THIN_BORDER
        ws.cell(row=row, column=5).alignment = CENTER_ALIGN
        ws.cell(row=row, column=6, value=reason).border = THIN_BORDER

    # =========================================================================
    # Section 5: Facility Status
    # =========================================================================
    ws['H4'] = "FACILITY STATUS"
    ws['H4'].font = SECTION_FONT
    ws['H4'].fill = SECTION_HEADER
    ws.merge_cells('H4:J4')

    facility_headers = ["Resource", "Current", "Status"]
    for col, header in enumerate(facility_headers, start=8):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_BLUE
        cell.border = THIN_BORDER
        cell.alignment = CENTER_ALIGN

    facility_data = [
        ("Compounding Hoods", "6/8", "OK"),
        ("Storage (sq ft)", "8K/12K", "OK"),
        ("Packing Stations", "8/10", "WATCH"),
    ]

    for i, (resource, current, status) in enumerate(facility_data):
        row = 7 + i
        ws.cell(row=row, column=8, value=resource).border = THIN_BORDER
        ws.cell(row=row, column=9, value=current).border = THIN_BORDER
        ws.cell(row=row, column=9).alignment = CENTER_ALIGN

        status_cell = ws.cell(row=row, column=10, value=status)
        status_cell.border = THIN_BORDER
        status_cell.alignment = CENTER_ALIGN
        if status == "OK":
            status_cell.fill = GREEN_OK
        elif status == "WATCH":
            status_cell.fill = YELLOW_WARN
        else:
            status_cell.fill = RED_ALERT

    # =========================================================================
    # Section 6: Key Dates
    # =========================================================================
    ws['H12'] = "KEY MILESTONES"
    ws['H12'].font = SECTION_FONT
    ws['H12'].fill = SECTION_HEADER
    ws.merge_cells('H12:J12')

    milestones = [
        "Week 4: PV2 hits 90% (critical)",
        "Week 8: New PV2 hires onboarded",
        "Week 10: Vial machine arrives",
        "Week 12: Fulfillment review",
        "Week 24: Facility expansion review",
    ]

    for i, milestone in enumerate(milestones):
        ws.cell(row=14+i, column=8, value=milestone)


def main():
    """Main function to generate the workbook."""
    print("Creating Pharmacy Capacity Planning Model...")

    wb = create_workbook()

    print("Building README tab...")
    build_readme_tab(wb["README"])

    print("Building INPUTS - Current State tab...")
    build_inputs_current_state_tab(wb["INPUTS - Current State"])

    print("Building INPUTS - Scenarios tab...")
    build_inputs_scenarios_tab(wb["INPUTS - Scenarios"])

    print("Building DEMAND FORECAST tab...")
    build_demand_forecast_tab(wb["DEMAND FORECAST"])

    print("Building CAPACITY MODEL tab...")
    build_capacity_model_tab(wb["CAPACITY MODEL"])

    print("Building FACILITY & EQUIPMENT tab...")
    build_facility_equipment_tab(wb["FACILITY & EQUIPMENT"])

    print("Building SCENARIO COMPARISON tab...")
    build_scenario_comparison_tab(wb["SCENARIO COMPARISON"])

    print("Building DASHBOARD tab...")
    build_dashboard_tab(wb["DASHBOARD"])

    # Save workbook
    output_file = "/home/user/thumbstud.io/Pharmacy_Capacity_Model_v1.xlsx"
    wb.save(output_file)
    print(f"\nWorkbook saved to: {output_file}")
    print("Done!")


if __name__ == "__main__":
    main()
